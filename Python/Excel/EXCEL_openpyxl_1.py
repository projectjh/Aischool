from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("diary", 0)    # 0이면 첫번째 위치에 삽입, 생략시 마지막 위치에 삽입

data = [('홍길동', 80, 70, 90), ('이기자', 90, 60, 80), ('강기자', 80, 80, 70)]

r = 1   # 첫번째 행이 1부터 시작
c = 1   # 첫번째 열이 1부터 시작

# 데이터 넣기 - 방법1
for irum, kor, eng, math in data:       # data는 튜플로 왼쪽과같이 작성 시 unpacking 발생하며 해당 변수에 값을 넣을 수 있다.
    ws.cell(row=r, column=c).value = irum
    ws.cell(row=r, column=c+1).value = kor
    ws.cell(row=r, column=c+2).value = eng
    ws.cell(row=r, column=c+3).value = math
    r += 1

"""
# 데이터 넣기 - 방법2
for irum, kor, eng, math in data:
    ws['A' + str(r)].value =  irum
    ws['B' + str(r)].value =  kor
    ws['C' + str(r)].value =  eng
    ws['D' + str(r)].value =  math
    r += 1
    
# 데이터 넣기 - 방법3
columChar = 'A'
for irum, kor, eng, math in data:
    ws[columnChar + str(r)].value = irum
    # ws[columnChar + str(r)].value = irum.encode(encoding='utf_8', errors='ignore')
    ws[chr(ord(columnChar)+1) + str(r)].value = kor     # ord(A)+1은 A문자의 다음 값을 나타냄 => B
    ws[chr(ord(columnChar)+2) + str(r)].value = eng
    ws[chr(ord(columnChar)+3) + str(r)].value = math
    r += 1
    
# 데이터 넣기 - 방법4
columChar = 'A'
for val in data:
    for i in range(0, 4):
        ws[chr(ord(columnChar)+i) + str(r)].value = val[i]
        
    r += 1
"""

ws.cell(row=r, column=1).value = '합계'
ws.cell(row=r, column=2).value = '=sum(B1:B3)'
ws.cell(row=r, column=3).value = '=sum(C1:C3)'
ws.cell(row=r, column=4).value = '=sum(D1:D3)'

wb.save("openpyxl2.xlsx")
wb.close()
