# https://openpyxl.readthedocs.io 사이트 참조
# 설치 : pip install openpyxl

from openpyxl import Workbook, load_workbook

wb = Workbook()     # Workbook()을 통해 새로운 엑셀 문서 생성 => 하나 이상의 Worksheet 로 생성 (cf. 워크 시트 추가는 Workbook.create_sheet())
ws = wb.active      # 활성화 되어 있는 시트 선택

ws.title = "sample"     # worksheet 이름 변경
ws['B2'] = 42
ws.append([1, 2, 3])    # append() => 현재 값이 있는 마지막 셀의 다음 행에 차례로 추가됨
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([1, 2, 3])
wb.save("openpyxl1.xlsx")
wb.close()              # 파일 다룰때는 생성 후 항상 닫아주는 작업이 필요함

wb = load_workbook(filename='openpyxl1.xlsx')   # load_workbook() => 이미 만들어져있는 파일을 불러올때 사용
ws = wb.active
ws['A1'] = 42                           # 셀 선택 방법1
ws.cell(row=1, column=3).value = 333    # 셀 선택 방법2
ws.append(['aaa', 'bbb', 'ccc'])        # 새로운 행에 추가

print(ws['A1'].value)
print(ws['A2'].value)   # 저장된 값이 없으면 None으로 출력

ws2 = wb['sample']  # sample 시트 선택
print(ws['A3'].value, ws2['B3'].value, ws2['C3'].value)
print(ws['A4'].value, ws2['B4'].value, ws2['C4'].value)
print(ws['A5'].value, ws2['B5'].value, ws2['C5'].value)

wb.save("openpyxl1.xlsx")
wb.close()  # 여기도 닫아주는 게 원칙
