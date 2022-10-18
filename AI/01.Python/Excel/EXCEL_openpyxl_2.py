from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active
ws = wb.create_sheet('chart', 0)

# ws.merge_cells('A1:D1')   # 셀 병합
ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
ws['A1'] = "성적표"
ws['A1'].font = Font(name="Hahmlet Light", size=15, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.append(['이름', '국어', '영어', '수학'])
ws.append(['홍길동', 60, 70, 60])
ws.append(['이기자', 88, 77, 89])
ws.append(['김기자', 55, 66, 77])

wb.save('openpyxl_chart.xlsx')  # 데이터 입력 후 1차 저장

barchart = BarChart()
barchart.title = "성적표"
barchart.x_axis.title = "이름"    # x축 타이틀
barchart.y_axis.title = "점수"    # y축 타이틀

data = Reference(ws, min_col=2, max_col=4, min_row=2, max_row=5)    # 차트 만들때 사용될 데이터 범위 지정
                # chart sheet에서 2열-4열, 2행-5행 데이터를 사용하겠다는 의미
cate = Reference(ws, min_col=1, min_row=3, max_row=5)               # 데이터 카테고리 범위 지정
                # 1열의 3행부터 5행까지 카테고리로 사용하겠다는 의미
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(cate)
barchart.shape = 1

ws.add_chart(barchart, 'F1')    # 'F1'은 차트가 만들어질 위치
wb.save('openpyxl_chart.xlsx')

wb.close()
